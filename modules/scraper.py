
# =========================================
# VERSION CORREGIDA V2
# =========================================
# Ajustes:
# - Datos abiertos fuera del loop de anexos
# - Excel una vez por proveedor
# - Mantiene extracción de emails
# - Mantiene extracción de teléfonos
# - Mantiene representante legal
# =========================================

from playwright.sync_api import sync_playwright
import re
import os
import pandas as pd
from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook
from PIL import Image
import pytesseract
from bs4 import BeautifulSoup
import requests

from modules.email_sender import enviar_correo
import time


todos_los_datos = []


# =========================================
# EXTRAER TEXTO ARCHIVOS
# =========================================

def extraer_texto_archivo(ruta):

    texto = ""

    try:

        extension = os.path.splitext(ruta)[1].lower()

        if extension == ".pdf":

            reader = PdfReader(ruta)

            for page in reader.pages:

                contenido = page.extract_text()

                if contenido:
                    texto += contenido + "\n"

            # ====================================
            # SI NO EXTRAJO TEXTO -> OCR
            # ====================================

            if len(texto.strip()) < 20:

                try:

                    import fitz

                    pdf = fitz.open(ruta)

                    for pagina in pdf:
                        pix = pagina.get_pixmap()

                        img_path = "data/temp/temp_pdf.png"

                        pix.save(img_path)

                        img = Image.open(img_path)

                        texto += pytesseract.image_to_string(
                            img,
                        )

                except Exception as e:

                    print("ERROR OCR PDF")
                    print(e)

        elif extension == ".docx":

            doc = Document(ruta)

            for p in doc.paragraphs:
                texto += p.text + "\n"

        elif extension in [".xlsx", ".xlsm", ".xls"]:

            wb = load_workbook(ruta, data_only=True)

            for ws in wb.worksheets:

                for row in ws.iter_rows(values_only=True):

                    texto += " ".join(
                        [str(x) for x in row if x]
                    ) + "\n"

        elif extension in [".png", ".jpg", ".jpeg"]:

            img = Image.open(ruta)

            texto += pytesseract.image_to_string(img)

        elif extension == ".txt":

            with open(ruta, "r", encoding="utf-8", errors="ignore") as f:
                texto = f.read()

    except Exception as e:

        print("\nERROR LEYENDO ARCHIVO")
        print(e)

    return texto


# =========================================
# SCRAPER PRINCIPAL
# =========================================

def abrir_licitacion(
        codigo_licitacion,
        empresa,
        rut,
        organismo
):

    os.makedirs("data/anexos", exist_ok=True)
    os.makedirs("data/output", exist_ok=True)
    os.makedirs("data/temp", exist_ok=True)

    datos = {

        "licitacion": codigo_licitacion,
        "empresa": empresa,
        "rut_empresa": rut,
        "organismo": organismo,

        "representante": "NO ENCONTRADO",

        "reclamos": "",
        "monto_adjudicado": "",

        "responsable_pago": "",
        "email_pago": "",
        "telefono_pago": "",

        "encargado_contrato": "",
        "email_contrato": "",
        "telefono_contrato": "",

        "monto_transado_anual": "",
        "ordenes_compra_anual": "",

        # ====================================
        # FICHA PROVEEDOR
        # ====================================

        "estado_habilidad": "",
        "motivo_inhabilidad": "",

        "socios": "",

        "ordenes_aceptadas": "",
        "sanciones": "",

    }

    with sync_playwright() as p:

        browser = p.chromium.launch(
            headless=False
        )

        context = browser.new_context()

        page = context.new_page()

        try:

            print("\nBUSCANDO LICITACION:")
            print(codigo_licitacion)

            page.goto(
                "https://www.mercadopublico.cl/Home/BusquedaLicitacion",
                wait_until="domcontentloaded",
                timeout=120000
            )

            page.wait_for_selector("#form-iframe")

            frame = page.frame_locator("#form-iframe")

            frame.locator("#textoBusqueda").wait_for()

            frame.locator("#textoBusqueda").fill(
                codigo_licitacion
            )

            frame.locator("#btnBuscar").click()

            page.wait_for_timeout(5000)

            html = page.frame(name="form-iframe").content()

            with open(
                "data/temp/resultados_iframe.html",
                "w",
                encoding="utf-8"
            ) as f:

                f.write(html)

            url_ficha = (
                "https://www.mercadopublico.cl/Procurement/Modules/RFB/"
                f"DetailsAcquisition.aspx?idlicitacion={codigo_licitacion}"
            )

            print("\\nABRIENDO:")
            print(url_ficha)

            print("\nABRIENDO:")
            print(url_ficha)

            page.goto(
                url_ficha,
                wait_until="domcontentloaded",
                timeout=120000
            )

            page.wait_for_timeout(5000)

            html_ficha = page.content()

            # =========================================
            # EXTRAER DATOS RESPONSABLES
            # =========================================

            soup = BeautifulSoup(html_ficha, "html.parser")

            texto_ficha = soup.get_text(" ", strip=True)

            texto_ficha = re.sub(r'\s+', ' ', texto_ficha)

            # ====================================
            # RESPONSABLE PAGO
            # ====================================

            match = re.search(
                r'Nombre de responsable de pago:\s*(.*?)\s*(?:e-mail de responsable de pago:)',
                texto_ficha,
                re.I
            )

            if match:
                datos["responsable_pago"] = match.group(1).strip()

            # ====================================
            # EMAIL PAGO
            # ====================================

            match = re.search(
                r'e-mail de responsable de pago:\s*([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]+)',
                texto_ficha,
                re.I
            )

            if match:
                datos["email_pago"] = match.group(1).strip()

            # ====================================
            # TELEFONO PAGO
            # ====================================

            bloque_pago = re.search(

                r'Nombre de responsable de pago:.*?'
                r'e-mail de responsable de pago:.*?'
                r'([+]?\d[\d\-\s]{7,20})',

                texto_ficha,
                re.I
            )

            if bloque_pago:
                telefono = re.sub(
                    r'\D',
                    '',
                    bloque_pago.group(1)
                )

                datos["telefono_pago"] = telefono

            # ====================================
            # ENCARGADO CONTRATO
            # ====================================

            match = re.search(
                r'Nombre de responsable de contrato:\s*(.*?)\s*(?:e-mail de responsable de contrato:)',
                texto_ficha,
                re.I
            )

            if match:
                datos["encargado_contrato"] = match.group(1).strip()

            # ====================================
            # EMAIL CONTRATO
            # ====================================

            match = re.search(
                r'e-mail de responsable de contrato:\s*([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]+)',
                texto_ficha,
                re.I
            )

            if match:
                datos["email_contrato"] = match.group(1).strip()

            # ====================================
            # TELEFONO CONTRATO
            # ====================================

            bloque_contrato = re.search(

                r'Nombre de responsable de contrato:.*?'
                r'e-mail de responsable de contrato:.*?'
                r'([+]?\d[\d\-\s]{7,20})',

                texto_ficha,
                re.I
            )

            if bloque_contrato:
                telefono = re.sub(
                    r'\D',
                    '',
                    bloque_contrato.group(1)
                )

                datos["telefono_contrato"] = telefono

            # =========================================
            # CUADRO OFERTAS
            # =========================================

            page.locator("#imgCuadroOferta").click(force=True)

            page.wait_for_timeout(5000)

            frame_ofertas = None

            for fr in page.frames:

                if "SupplySummary.aspx" in fr.url:
                    frame_ofertas = fr

            if frame_ofertas is None:

                print("NO SE ENCONTRO FRAME")
                browser.close()
                return

            frame_ofertas.wait_for_timeout(5000)

            anexos_visitados = set()

            # =========================================
            # CARPETA LICITACION
            # =========================================

            carpeta_licitacion = os.path.join(
                "data/anexos",
                codigo_licitacion
            )

            os.makedirs(
                carpeta_licitacion,
                exist_ok=True
            )

            filas = frame_ofertas.locator(
                "#grdSupplies tbody tr"
            )

            total_filas = filas.count()

            print("\nTOTAL FILAS:")
            print(total_filas)

            for fila_idx in range(1, total_filas):

                try:

                    fila = filas.nth(fila_idx)

                    columnas = fila.locator("td")

                    total_columnas = columnas.count()

                    if total_columnas < 3:
                        continue

                    rut_proveedor = columnas.nth(0).inner_text().strip()

                    proveedor = columnas.nth(1).inner_text().strip()

                    if "declaración jurada" in proveedor.lower():
                        continue

                    monto_oferta = ""

                    if total_columnas >= 4:
                        monto_oferta = columnas.nth(3).inner_text().strip()

                    botones = fila.locator("[onclick]")

                    total_botones = botones.count()

                    popup_proveedor_abierto = False

                    anexos_descargados = []

                    # =========================================
                    # VARIABLES ACUMULADORAS
                    # =========================================

                    emails_acumulados = []
                    telefonos_acumulados = []
                    representante_acumulado = ""

                    datos_proveedor = {

                        "licitacion": codigo_licitacion,
                        "empresa": proveedor,
                        "rut_empresa": rut_proveedor,
                        "representante": "NO ENCONTRADO",

                        "organismo": organismo,
                        "reclamos": datos.get("reclamos", ""),
                        "monto_adjudicado": monto_oferta,

                        "responsable_pago": datos.get("responsable_pago", ""),
                        "email_pago": datos.get("email_pago", ""),
                        "telefono_pago": datos.get("telefono_pago", ""),

                        "encargado_contrato": datos.get("encargado_contrato", ""),
                        "email_contrato": datos.get("email_contrato", ""),
                        "telefono_contrato": datos.get("telefono_contrato", ""),

                        "monto_transado_anual": "",
                        "ordenes_compra_anual": ""
                    }

                    for b in range(total_botones):

                        try:

                            boton = botones.nth(b)

                            title = boton.get_attribute("title")
                            onclick = boton.get_attribute("onclick")

                            texto_boton = str(title).lower() if title else ""

                            # =========================================
                            # DATOS PROVEEDOR
                            # =========================================

                            if (
                                onclick
                                and "InformationProvider.aspx" in onclick
                                and not popup_proveedor_abierto
                            ):

                                match = re.search(
                                    r"InformationProvider\\.aspx\\?enc=[^']+",
                                    onclick
                                )

                                if match:

                                    url_proveedor = (
                                        "https://www.mercadopublico.cl/BID/Modules/PopUps/"
                                        + match.group(0)
                                    )

                                    popup_info = browser.new_page()

                                    popup_info.goto(
                                        url_proveedor,
                                        wait_until="networkidle",
                                        timeout=120000
                                    )

                                    texto_popup = popup_info.inner_text("body")

                                    emails_popup = re.findall(
                                        r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\\.[a-zA-Z]+',
                                        texto_popup
                                    )

                                    telefonos_popup = re.findall(
                                        r'\\+?\\d[\\d\\s\\-]{8,15}',
                                        texto_popup
                                    )

                            # =========================================
                            # ANEXOS
                            # =========================================

                            archivos_objetivo = [

                                "anexo 1",
                                "anexo1",
                                "anexo_1",
                                "anexo n°1",
                                "anexo nº1",
                                "anexon°1",
                                "anexo i",

                                "anexo 2",
                                "anexo2",
                                "anexo_2",
                                "anexo n°2",
                                "anexo nº2",

                                "anexo a",
                                "anexoa",
                                "anexo_a",

                                "formulario 1",
                                "formulario1",
                                "formulario_1",
                                "formulario n1",
                                "formulario n°1",
                                "formulario_n°1",

                                "identificacion del oferente",
                                "identificación del oferente",

                                "identificacion del proveedor",
                                "identificación del proveedor",

                                "antecedentes laborales",
                                "antecedentes previsionales",

                                "antecedentes del oferente",
                                "antecedentes del proveedor",

                                "f30",
                                "f30-1"
                            ]

                            if onclick and "ViewBidAttachment" in onclick:

                                if not any(
                                        x in texto_boton
                                        for x in [
                                            "administrativos",
                                            "tecnicos",
                                            "técnicos",
                                            "economicos",
                                            "económicos"
                                        ]
                                ):
                                    continue

                                print("\nANEXO DETECTADO")

                                try:

                                        match = re.search(
                                            r"ViewBidAttachment\.aspx\?enc=[^']+",
                                            onclick
                                        )

                                        if not match:
                                            continue

                                        url_anexo = (
                                                "https://www.mercadopublico.cl/BID/Modules/POPUPS/"
                                                + match.group(0)
                                        )

                                        if url_anexo in anexos_visitados:
                                            continue

                                        anexos_visitados.add(url_anexo)

                                        print(url_anexo)

                                        popup = browser.new_page()

                                        popup.goto(
                                            url_anexo,
                                            wait_until="domcontentloaded",
                                            timeout=120000
                                        )

                                        popup.wait_for_timeout(5000)

                                        paginas = popup.locator("a")

                                        total_paginas = paginas.count()

                                        for p in range(total_paginas):

                                            try:

                                                texto_pagina = paginas.nth(p).inner_text().strip()

                                                if texto_pagina.isdigit():
                                                    paginas.nth(p).click()

                                                    popup.wait_for_timeout(3000)

                                            except:
                                                pass

                                        # =========================================
                                        # BOTONES VER
                                        # =========================================

                                        # ====================================
                                        # BUSCAR LINKS REALES
                                        # ====================================

                                        # ====================================
                                        # BOTONES DESCARGA REALES
                                        # ====================================

                                        popup.wait_for_timeout(5000)

                                        # =========================================
                                        # BOTONES VER
                                        # =========================================

                                        # =========================================
                                        # FILAS REALES DE ANEXOS
                                        # =========================================

                                        filas_anexos = popup.locator("table tr")

                                        total_anexos = filas_anexos.count()

                                        print("\nTOTAL ANEXOS:")
                                        print(total_anexos)

                                        archivos_objetivo = [

                                            "anexo 1",
                                            "anexo1",
                                            "anexo_1",
                                            "anexo n°1",
                                            "anexo nº1",
                                            "anexon°1",

                                            "anexo a",
                                            "anexoa",
                                            "anexon_a",

                                            "formulario 1",
                                            "formulario1",
                                            "formulario_1",
                                            "formulario n1",
                                            "formulario n°1",
                                            "formulario_n°1",

                                            "anexo 2",
                                            "anexo2",
                                            "anexo_2",
                                            "anexo n°2",
                                            "anexo nº2",

                                            "identificacion del oferente",
                                            "identificación del oferente",

                                            "identificacion del proveedor",
                                            "identificación del proveedor",

                                            "antecedentes laborales",
                                            "antecedentes previsionales",

                                            "antecedentes del oferente",
                                            "antecedentes del proveedor",

                                            "f30",
                                            "f30-1"
                                        ]

                                        for j in range(total_anexos):

                                            try:

                                                fila_anexo = filas_anexos.nth(j)

                                                celdas = fila_anexo.locator("td")

                                                if celdas.count() < 2:
                                                    continue

                                                texto_fila = celdas.nth(1).inner_text(timeout=5000).lower().strip()

                                                if texto_fila in ["1", "2", "3", "4", "5"]:
                                                    continue

                                                if "generar nuevo código" in texto_fila:
                                                    continue

                                                # =========================================
                                                # NORMALIZAR NOMBRE ARCHIVO
                                                # =========================================

                                                texto_normalizado = texto_fila.lower()

                                                # reemplazar separadores
                                                texto_normalizado = texto_normalizado.replace("_", " ")
                                                texto_normalizado = texto_normalizado.replace("-", " ")

                                                # normalizar N°
                                                texto_normalizado = texto_normalizado.replace("n°", "n")
                                                texto_normalizado = texto_normalizado.replace("nº", "n")

                                                # quitar espacios dobles
                                                texto_normalizado = re.sub(r'\s+', ' ', texto_normalizado)

                                                texto_normalizado = texto_normalizado.strip()

                                                # reemplazos generales
                                                texto_normalizado = (
                                                    texto_normalizado
                                                    .replace("_", "")
                                                    .replace("-", "")
                                                    .replace(" ", "")
                                                    .replace(".", "")
                                                    .replace("º", "°")
                                                    .replace("n°", "n")
                                                    .replace("nº", "n")
                                                    .replace("°", "")
                                                    .replace("ñ", "n")
                                                    .replace("á", "a")
                                                    .replace("é", "e")
                                                    .replace("í", "i")
                                                    .replace("ó", "o")
                                                    .replace("ú", "u")
                                                )

                                                # eliminar caracteres raros
                                                texto_normalizado = re.sub(
                                                    r'[^a-z0-9]',
                                                    '',
                                                    texto_normalizado
                                                )

                                                # =========================================
                                                # PALABRAS OBJETIVO NORMALIZADAS
                                                # =========================================

                                                archivos_objetivo = [

                                                    "anexo1",
                                                    "anexon1",
                                                    "anexoa",

                                                    "anexo2",
                                                    "anexon2",

                                                    "formulario1",
                                                    "formularion1",

                                                    "identificaciondeloferente",
                                                    "identificaciondelproveedor",

                                                    "antecedenteslaborales",
                                                    "antecedentesprevisionales",

                                                    "antecedentesdeloferente",
                                                    "antecedentesdelproveedor",

                                                    "f30",
                                                    "f301"
                                                ]

                                                # =========================================
                                                # BUSQUEDA FLEXIBLE
                                                # =========================================

                                                archivo_valido = False

                                                palabras_clave = [

                                                    "anexo1",
                                                    "anexon1",
                                                    "anexon°1",
                                                    "anexonº1",

                                                    "formulario1",
                                                    "formularion1",
                                                    "formularion°1",
                                                    "formularionº1",

                                                    "identificaciondeloferente",
                                                    "identificaciondelproveedor",
                                                    "identificacion",

                                                    "anexos1",
                                                    "anexo_1",
                                                    "anexo-1",

                                                    "formularion1identificacion",
                                                ]

                                                for palabra in palabras_clave:

                                                    if palabra in texto_normalizado:
                                                        archivo_valido = True
                                                        break

                                                # ====================================
                                                # SI NO ES ARCHIVO OBJETIVO
                                                # ====================================

                                                if not archivo_valido:
                                                    continue

                                                print("\nARCHIVO OBJETIVO ENCONTRADO")
                                                print(texto_fila)

                                                boton_ver = fila_anexo.locator(
                                                    "input[type='image']"
                                                )

                                                if boton_ver.count() == 0:
                                                    continue

                                                with popup.expect_download(timeout=120000) as download_info:

                                                    boton_ver.nth(0).click(force=True)


                                                    download = download_info.value

                                                    nombre_archivo = download.suggested_filename

                                                    nombre_archivo = nombre_archivo.strip().lower()

                                                    if (
                                                            nombre_archivo in ["1", "2"]
                                                            or "generar nuevo código" in nombre_archivo
                                                            or len(nombre_archivo) < 5
                                                    ):
                                                        continue

                                                    ruta_guardado = os.path.join(
                                                        carpeta_licitacion,
                                                        nombre_archivo
                                                    )

                                                    download.save_as(ruta_guardado)

                                                    if ruta_guardado not in anexos_descargados:
                                                        anexos_descargados.append(ruta_guardado)

                                                    print("\nARCHIVO DESCARGADO:")
                                                    print(ruta_guardado)

                                            except Exception as e:

                                                print("\nERROR DESCARGANDO:")
                                                print(e)

                                        try:
                                            popup.close()
                                        except:
                                            pass

                                        # =========================================
                                        # LEER ANEXOS DESCARGADOS
                                        # =========================================

                                        texto_total = ""

                                        for ruta_anexo in anexos_descargados:

                                            try:

                                                texto = extraer_texto_archivo(ruta_anexo)

                                                texto_total += "\n" + texto

                                            except Exception as e:

                                                print("\nERROR LEYENDO ANEXO")
                                                print(e)

                                        # =========================================
                                        # EXTRAER EMAILS
                                        # =========================================

                                        emails_anexo = re.findall(
                                            r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]+',
                                            texto_total
                                        )

                                        telefonos_anexo = re.findall(
                                            r'(?:\+?56[\s\-]?)?(?:0?2|0?9)?[\s\-]?\d{4,5}[\s\-]?\d{4}',
                                            texto_total
                                        )

                                        # =========================================
                                        # EMAILS
                                        # =========================================

                                        emails_unicos = []

                                        for em in emails_anexo:

                                            em = em.lower().strip()

                                            if em not in emails_unicos:
                                                emails_unicos.append(em)

                                        if not emails_unicos:

                                            datos_proveedor["email_1"] = "NO ENCONTRADO"

                                        else:

                                            for i, em in enumerate(emails_unicos, start=1):
                                                if em not in emails_acumulados:
                                                    emails_acumulados.append(em)

                                        # =========================================
                                        # TELEFONOS
                                        # =========================================

                                        telefonos_unicos = []

                                        for tel in telefonos_anexo:

                                            tel = re.sub(r"\D", "", tel)

                                            if tel.startswith("56"):
                                                tel = tel[2:]

                                            if len(tel) == 8:
                                                if tel.startswith("2"):
                                                    tel = tel
                                                else:
                                                    tel = "9" + tel

                                            tel = tel.strip()

                                            if tel not in telefonos_unicos:
                                                telefonos_unicos.append(tel)

                                        if not telefonos_unicos:

                                            datos_proveedor["telefono_1"] = "NO ENCONTRADO"

                                        else:

                                            for i, tel in enumerate(telefonos_unicos, start=1):
                                                if tel not in telefonos_acumulados:
                                                    telefonos_acumulados.append(tel)

                                        # =========================================
                                        # REPRESENTANTE
                                        # =========================================

                                        # =========================================
                                        # REPRESENTANTE LEGAL
                                        # =========================================

                                        representante_final = "NO ENCONTRADO"

                                        texto_limpio = texto_total.upper()

                                        texto_limpio = re.sub(r'\s+', ' ', texto_limpio)

                                        # ====================================
                                        # PATRONES AVANZADOS
                                        # ====================================

                                        # ====================================
                                        # BUSCAR BLOQUE REPRESENTANTE LEGAL
                                        # ====================================

                                        if "IDENTIFICACION DEL REPRESENTANTE LEGAL" in texto_limpio:

                                            try:

                                                bloque_rep = texto_limpio.split(
                                                    "IDENTIFICACION DEL REPRESENTANTE LEGAL"
                                                )[1]

                                                match_nombre = re.search(
                                                    r'NOMBRE\s*[:\-]?\s*([A-ZÁÉÍÓÚÑ ]{6,80})',
                                                    bloque_rep
                                                )

                                                if match_nombre:

                                                    nombre_directo = match_nombre.group(1).strip()

                                                    nombre_directo = re.sub(
                                                        r'\s+',
                                                        ' ',
                                                        nombre_directo
                                                    )

                                                    basura_rep = [

                                                        "RAZON SOCIAL",
                                                        "RUT",
                                                        "DIRECCION",
                                                        "DOMICILIO",
                                                        "TELEFONO",
                                                        "EMAIL",
                                                        "CORREO",
                                                        "CELULAR",
                                                        "COMUNA",
                                                        "CIUDAD",
                                                        "FIRMA"
                                                    ]

                                                    valido = True

                                                    for b in basura_rep:

                                                        if b in nombre_directo:
                                                            valido = False
                                                            break

                                                    if valido and len(nombre_directo.split()) >= 2:
                                                        representante_final = nombre_directo.title()

                                            except:
                                                pass

                                        patrones_representante = [

                                            r'NOMBRE\s+REPRESENTANTE\s+LEGAL\s*[:\-]?\s*([A-ZÁÉÍÓÚÑ ]{6,80})',

                                            r'NOMBRE\s+DEL\s+REPRESENTANTE\s+LEGAL\s*[:\-]?\s*([A-ZÁÉÍÓÚÑ ]{6,80})',

                                            r'REPRESENTANTE\s+LEGAL\s*[:\-]?\s*([A-ZÁÉÍÓÚÑ ]{6,80})',

                                            r'REP\.?\s*LEGAL\s*[:\-]?\s*([A-ZÁÉÍÓÚÑ ]{6,80})',

                                            r'NOMBRE\s+REPRESENTANTE\s+LEGAL\s*[:\-]?\s*([A-ZÁÉÍÓÚÑ ]{6,80})',

                                            r'NOMBRE\s+DEL\s+REPRESENTANTE\s+LEGAL\s*[:\-]?\s*([A-ZÁÉÍÓÚÑ ]{6,80})',

                                            r'REPRESENTANTE\s+LEGAL\s*[:\-]?\s*([A-ZÁÉÍÓÚÑ ]{6,80})',

                                            r'REP\.?\s*LEGAL\s*[:\-]?\s*([A-ZÁÉÍÓÚÑ ]{6,80})',

                                            r'IDENTIFICACION\s+DEL\s+REPRESENTANTE\s+LEGAL.*?NOMBRE\s*[:\-]?\s*([A-ZÁÉÍÓÚÑ ]{6,80})'
                                        ]

                                        for patron in patrones_representante:

                                            resultado = re.search(
                                                patron,
                                                texto_limpio,
                                                re.I
                                            )

                                            if resultado:

                                                nombre = resultado.group(1).strip()

                                                # ====================================
                                                # CORTAR BASURA
                                                # ====================================

                                                basura = [

                                                    "RUT",
                                                    "DIRECCION",
                                                    "DOMICILIO",
                                                    "TELEFONO",
                                                    "EMAIL",
                                                    "CORREO",
                                                    "CELULAR",
                                                    "COMUNA",
                                                    "CIUDAD",
                                                    "ANTECEDENTES",
                                                    "CERTIFICA",
                                                    "RAZON SOCIAL",
                                                    "NOMBRE",
                                                    "CARGO",
                                                    "ABOGADO",
                                                    "FAX",
                                                    "E MAIL",
                                                    "REPRESENTANTE",
                                                    "LEGAL DEL OFERENTE"
                                                ]

                                                cortar = len(nombre)

                                                for b in basura:

                                                    pos = nombre.find(b)

                                                    if pos != -1:
                                                        cortar = min(cortar, pos)

                                                nombre = nombre[:cortar].strip()

                                                nombre = re.sub(r'\s+', ' ', nombre)

                                                # ====================================
                                                # VALIDAR
                                                # ====================================

                                                palabras = nombre.split()

                                                # ====================================
                                                # DESCARTAR BASURA
                                                # ====================================

                                                basura_extra = [

                                                    "LOS ANGELES",
                                                    "FIRMA",
                                                    "FIRMADO",
                                                    "NOMBRE Y FIRMA",
                                                    "RAZON SOCIAL",
                                                    "RUT",
                                                    "DIRECCION",
                                                    "DOMICILIO",
                                                    "TELEFONO",
                                                    "EMAIL",
                                                    "CORREO",
                                                    "CELULAR",
                                                    "COMUNA",
                                                    "CIUDAD",
                                                    "MUNICIPALIDAD",
                                                    "CHILE",
                                                    "PROVEEDOR",
                                                    "OFERENTE",
                                                    "REPRESENTANTE",
                                                    "LEGAL",
                                                    "FONO"
                                                ]

                                                if any(
                                                        b in nombre
                                                        for b in basura_extra
                                                ):
                                                    continue

                                                if (
                                                        len(palabras) >= 2
                                                        and len(nombre) >= 8
                                                ):

                                                    if len(nombre.split()) < 2:
                                                        continue

                                                    if any(char.isdigit() for char in nombre):
                                                        continue

                                                    if "FIRMA" in nombre:
                                                        continue
                                                    representante_final = nombre.title()
                                                    break

                                        if (
                                                representante_final != "NO ENCONTRADO"
                                                and not representante_acumulado
                                        ):
                                            representante_acumulado = representante_final

                                except Exception as e:

                                    print("\nERROR ANEXO")
                                    print(e)

                        except Exception as e:

                            print("\nERROR BOTON")
                            print(e)

                    # =========================================
                    # CONSOLIDAR DATOS
                    # =========================================

                    if representante_acumulado:
                        datos_proveedor["representante"] = representante_acumulado

                    for i, em in enumerate(emails_acumulados, start=1):
                        datos_proveedor[f"email_{i}"] = em

                    for i, tel in enumerate(telefonos_acumulados, start=1):
                        datos_proveedor[f"telefono_{i}"] = tel

                    # =========================================
                    # DATOS ABIERTOS
                    # =========================================

                    page_datos = None

                    try:

                        if not rut_proveedor or len(rut_proveedor) < 8:
                            print("RUT INVALIDO")
                            continue

                        url_datos_abiertos = (
                            f"https://datos-abiertos.chilecompra.cl/"
                            f"organismos-proveedores/{rut_proveedor}"
                        )

                        page_datos = browser.new_page()

                        page_datos.goto(
                            url_datos_abiertos,
                            wait_until="domcontentloaded",
                            timeout=120000
                        )

                        page_datos.wait_for_timeout(5000)

                        # =========================================
                        # BUSCAR AÑOS DISPONIBLES
                        # =========================================

                        try:

                            page_datos.wait_for_timeout(5000)

                            años = ["2026", "2025", "2024", "2023", "2022"]

                            for año in años:

                                try:

                                    print(f"\nCAMBIANDO AÑO: {año}")

                                    # ====================================
                                    # ABRIR SELECTOR AÑO REAL
                                    # ====================================

                                    selector_anio = page_datos.locator(
                                        "button"
                                    ).filter(
                                        has_text=re.compile(r"20(22|23|24|25|26)")
                                    ).first

                                    if selector_anio.count() == 0:
                                        print("NO HAY DATOS ABIERTOS")
                                        break

                                    selector_anio.click(force=True)

                                    page_datos.wait_for_timeout(2000)

                                    page_datos.locator(
                                        f"text='{año}'"
                                    ).nth(0).click(force=True)

                                    page_datos.wait_for_timeout(5000)

                                    # ====================================
                                    # EXTRAER GRAFICO ACTUAL
                                    # ====================================

                                    texto_datos = page_datos.locator("body").inner_text()

                                    lineas = texto_datos.splitlines()

                                    monto_actual = 0
                                    orden_actual = 0

                                    for idx, linea in enumerate(lineas):

                                        linea_limpia = linea.strip().lower()

                                        if "$" in linea:

                                            numeros = re.findall(r'[\d\.\,]+', linea)

                                            if numeros:

                                                valor = numeros[0]

                                                valor = (
                                                    valor
                                                    .replace(".", "")
                                                    .replace(",", "")
                                                )

                                                try:
                                                    monto_actual = int(valor)
                                                    break

                                                except:
                                                    pass

                                    orden_actual = 0

                                    for idx, linea in enumerate(lineas):

                                        texto_linea = linea.strip().lower()

                                        # buscar linea del año actual
                                        if año in texto_linea:

                                            bloque = "\n".join(
                                                lineas[idx:idx + 8]
                                            )

                                            # buscar ordenes SOLO CERCA del año
                                            match_orden = re.search(
                                                r'(\d+)\s+órdenes',
                                                bloque,
                                                re.I
                                            )

                                            if match_orden:

                                                try:
                                                    orden_actual = int(
                                                        match_orden.group(1)
                                                    )

                                                    break

                                                except:
                                                    pass

                                    print("MONTO:", monto_actual)
                                    print("ORDENES:", orden_actual)

                                    datos_proveedor[f"monto_{año}"] = monto_actual
                                    datos_proveedor[f"ordenes_{año}"] = orden_actual

                                except Exception as e:

                                    print(f"ERROR AÑO {año}")
                                    print(e)

                            # =========================================
                            # FICHA PROVEEDOR MERCADO PUBLICO
                            # =========================================

                            try:

                                url_ficha_proveedor = (
                                    f"https://proveedor.mercadopublico.cl/"
                                    f"ficha/informacion-general/{rut_proveedor}"
                                )

                                page_ficha = browser.new_page()

                                page_ficha.goto(
                                    url_ficha_proveedor,
                                    wait_until="domcontentloaded",
                                    timeout=120000
                                )

                                page_ficha.wait_for_load_state("networkidle")

                                page_ficha.wait_for_timeout(8000)

                                try:

                                    page_ficha.wait_for_selector(
                                        "text=Socios y accionistas",
                                        timeout=15000
                                    )

                                except:
                                    pass

                                texto_ficha = page_ficha.locator("body").inner_text()

                                print("\n======================")
                                print("HTML FICHA PROVEEDOR")
                                print("======================")

                                html_debug = page_ficha.content()

                                with open(
                                        "data/temp/debug_socios.html",
                                        "w",
                                        encoding="utf-8"
                                ) as f:

                                    f.write(html_debug)

                                print("HTML GUARDADO")

                                try:

                                    page_ficha.get_by_role(
                                        "heading",
                                        name=re.compile("Socios y accionistas", re.I)
                                    ).click()

                                    page_ficha.wait_for_timeout(4000)

                                except Exception as e:

                                    print("NO SE PUDO ABRIR SOCIOS")
                                    print(e)

                                # ====================================
                                # SOCIOS Y ACCIONISTAS
                                # ====================================

                                try:

                                    socios_finales = []

                                    filas_socios = page_ficha.locator(
                                        "tbody tr"
                                    )

                                    print("TOTAL FILAS SOCIOS:")
                                    print(filas_socios.count())

                                    total_socios = filas_socios.count()

                                    for i in range(total_socios):

                                        try:

                                            fila = filas_socios.nth(i)

                                            celdas = fila.locator("td")

                                            if celdas.count() < 3:
                                                continue

                                            valores = []

                                            for c in range(celdas.count()):

                                                try:

                                                    valor = celdas.nth(c).inner_text().strip()

                                                    valores.append(valor)

                                                except:
                                                    pass

                                            print("VALORES BENEFICIARIO:")
                                            print(valores)

                                            nombre = ""
                                            porcentaje = ""

                                            for valor in valores:

                                                if "%" in valor:

                                                    porcentaje = valor

                                                elif (
                                                        valor
                                                        and len(valor) > 4
                                                        and "%" not in valor
                                                ):

                                                    nombre = valor

                                            if (
                                                    nombre
                                                    and "%" in porcentaje
                                                    and "Nombre" not in nombre
                                                    and len(nombre) > 4
                                            ):

                                                socio = f"{nombre} ({porcentaje})"

                                                if socio not in socios_finales:
                                                    socios_finales.append(socio)

                                        except:
                                            pass

                                    datos_proveedor["socios"] = " | ".join(socios_finales)

                                    print("SOCIOS:")
                                    print(datos_proveedor["socios"])

                                except Exception as e:

                                    print("ERROR SOCIOS")
                                    print(e)

                                # ====================================
                                # BENEFICIARIO FINAL
                                # ====================================

                                try:

                                    beneficiarios_finales = []

                                    tablas = page_ficha.locator("table")

                                    total_tablas = tablas.count()

                                    print("TOTAL TABLAS:")
                                    print(total_tablas)

                                    for t in range(total_tablas):

                                        try:

                                            tabla_benef = tablas.nth(t)

                                            filas = tabla_benef.locator("tbody tr")

                                            total_filas = filas.count()

                                            print("FILAS BENEFICIARIO:")
                                            print(total_filas)

                                            for i in range(total_filas):

                                                try:

                                                    fila = filas.nth(i)

                                                    celdas = fila.locator("td")

                                                    if celdas.count() < 3:
                                                        continue

                                                    valores = []

                                                    for c in range(celdas.count()):
                                                        valor = celdas.nth(c).inner_text().strip()

                                                        valores.append(valor)

                                                    print("VALORES BENEFICIARIO:")
                                                    print(valores)

                                                    nombre = ""
                                                    porcentaje = ""

                                                    # ====================================
                                                    # FORMATO REAL:
                                                    # ['HARRY...', '-', '100.00%']
                                                    # ====================================

                                                    if len(valores) >= 3:

                                                        posible_nombre = valores[0].strip()
                                                        posible_porcentaje = valores[2].strip()

                                                        if (
                                                                posible_nombre
                                                                and "%" in posible_porcentaje
                                                                and "nombre" not in posible_nombre.lower()
                                                        ):
                                                            nombre = posible_nombre
                                                            porcentaje = posible_porcentaje

                                                    print("NOMBRE BENEFICIARIO:", nombre)
                                                    print("PORCENTAJE BENEFICIARIO:", porcentaje)

                                                    if nombre and porcentaje:

                                                        beneficiario = (
                                                            f"{nombre} ({porcentaje})"
                                                        )

                                                        if beneficiario not in beneficiarios_finales:
                                                            beneficiarios_finales.append(
                                                                beneficiario
                                                            )

                                                            print("BENEFICIARIO ENCONTRADO:")
                                                            print(beneficiario)

                                                            break

                                                except:
                                                    pass

                                            # SI YA ENCONTRO BENEFICIARIO
                                            if beneficiarios_finales:
                                                break

                                        except:
                                            pass

                                    datos_proveedor["beneficiario_final"] = " | ".join(
                                        beneficiarios_finales
                                    )

                                    print("BENEFICIARIO FINAL:")
                                    print(datos_proveedor["beneficiario_final"])

                                except Exception as e:

                                    print("ERROR BENEFICIARIO")
                                    print(e)

                                # ====================================
                                # PESTAÑA HABILIDAD
                                # ====================================

                                try:

                                    page_ficha.get_by_role("tab", name="Habilidad").click()

                                    page_ficha.wait_for_timeout(4000)

                                    texto_habilidad = page_ficha.locator("body").inner_text()

                                    if "INHÁBIL" in texto_habilidad.upper():

                                        datos_proveedor["estado_habilidad"] = "INHÁBIL"

                                        motivos = []

                                        patrones = [

                                            "DOCUMENTACIÓN FALSA",
                                            "SUSPENSIÓN DEL REGISTRO DE PROVEEDORES",
                                            "COHECHO",
                                            "FINANCIAMIENTO DEL TERRORISMO",
                                            "LIQUIDACIÓN CONCURSAL",
                                            "DEUDAS TRIBUTARIAS",
                                            "SENTENCIAS INFORMADAS"
                                        ]

                                        for p in patrones:

                                            if p in texto_habilidad.upper():
                                                motivos.append(p)

                                        datos_proveedor["motivo_inhabilidad"] = " | ".join(motivos)

                                    else:

                                        datos_proveedor["estado_habilidad"] = "HÁBIL"

                                except Exception as e:

                                    print("ERROR HABILIDAD")
                                    print(e)

                                # ====================================
                                # PESTAÑA COMPORTAMIENTO
                                # ====================================

                                try:

                                    page_ficha.get_by_role("tab", name="Comportamiento").click()

                                    page_ficha.wait_for_timeout(4000)

                                    texto_comp = page_ficha.locator("body").inner_text()

                                    # ====================================
                                    # ORDENES ACEPTADAS
                                    # ====================================

                                    match_ordenes = re.search(
                                        r'(\d+)\s+Órdenes de compra aceptadas',
                                        texto_comp,
                                        re.I
                                    )

                                    if match_ordenes:
                                        datos_proveedor["ordenes_aceptadas"] = (
                                            match_ordenes.group(1)
                                        )

                                    # ====================================
                                    # SANCIONES
                                    # ====================================

                                    match_sanciones = re.search(
                                        r'(\d+)\s+Sanciones',
                                        texto_comp,
                                        re.I
                                    )

                                    if match_sanciones:
                                        datos_proveedor["sanciones"] = (
                                            match_sanciones.group(1)
                                        )

                                except Exception as e:

                                    print("ERROR COMPORTAMIENTO")
                                    print(e)

                                try:
                                    page_ficha.close()
                                except:
                                    pass

                            except Exception as e:

                                print("ERROR FICHA PROVEEDOR")
                                print(e)

                        except Exception as e:





                            print("ERROR DATOS ABIERTOS")
                            print(e)

                    finally:

                        try:
                            page_datos.close()
                        except:
                            pass

                        context.clear_cookies()

                    # =========================================
                    # GUARDAR UNICO
                    # =========================================

                    existe = False

                    for item in todos_los_datos:

                        if (
                                item["licitacion"] == datos_proveedor["licitacion"]
                                and item["rut_empresa"] == datos_proveedor["rut_empresa"]
                        ):
                            existe = True
                            break

                    if not existe:
                        todos_los_datos.append(datos_proveedor)

                    # =========================================
                    # HOJA 1
                    # SOLO DATOS LICITACION
                    # =========================================

                    datos_hoja1 = []

                    for fila in todos_los_datos:
                        datos_hoja1.append({

                            "licitacion": fila.get("licitacion", ""),
                            "empresa": fila.get("empresa", ""),
                            "rut_empresa": fila.get("rut_empresa", ""),

                            "representante": fila.get("representante", ""),

                            "organismo": fila.get("organismo", ""),

                            "monto_adjudicado": fila.get("monto_adjudicado", ""),

                            "responsable_pago": fila.get("responsable_pago", ""),
                            "email_pago": fila.get("email_pago", ""),
                            "telefono_pago": fila.get("telefono_pago", ""),

                            "encargado_contrato": fila.get("encargado_contrato", ""),
                            "email_contrato": fila.get("email_contrato", ""),
                            "telefono_contrato": fila.get("telefono_contrato", ""),

                            "email_1": fila.get("email_1", ""),
                            "telefono_1": fila.get("telefono_1", ""),

                            "email_2": fila.get("email_2", ""),
                            "telefono_2": fila.get("telefono_2", ""),
                        })

                    df_hoja1 = pd.DataFrame(datos_hoja1)

                    # =========================================
                    # HOJA 2
                    # ANALISIS PROVEEDOR
                    # =========================================

                    datos_hoja2 = []

                    for fila in todos_los_datos:
                        datos_hoja2.append({

                            "empresa": fila.get("empresa", ""),
                            "rut_empresa": fila.get("rut_empresa", ""),

                            "socios": fila.get("socios", ""),

                            "beneficiario_final": fila.get("beneficiario_final", ""),

                            "estado_habilidad": fila.get("estado_habilidad", ""),
                            "motivo_inhabilidad": fila.get("motivo_inhabilidad", ""),

                            "ordenes_aceptadas": fila.get("ordenes_aceptadas", ""),
                            "sanciones": fila.get("sanciones", ""),

                            "monto_2026": fila.get("monto_2026", 0),
                            "ordenes_2026": fila.get("ordenes_2026", 0),

                            "monto_2025": fila.get("monto_2025", 0),
                            "ordenes_2025": fila.get("ordenes_2025", 0),

                            "monto_2024": fila.get("monto_2024", 0),
                            "ordenes_2024": fila.get("ordenes_2024", 0),

                            "monto_2023": fila.get("monto_2023", 0),
                            "ordenes_2023": fila.get("ordenes_2023", 0),

                            "monto_2022": fila.get("monto_2022", 0),
                            "ordenes_2022": fila.get("ordenes_2022", 0),
                        })

                    df_hoja2 = pd.DataFrame(datos_hoja2)

                    # =========================================
                    # EXPORTAR MULTIHOJA
                    # =========================================

                    # =========================================
                    # EXPORTAR MULTIHOJA
                    # =========================================

                    ruta_excel = "data/output/proveedores.xlsx"

                    with pd.ExcelWriter(
                            ruta_excel,
                            engine="openpyxl"
                    ) as writer:

                        df_hoja1.to_excel(
                            writer,
                            sheet_name="proveedores",
                            index=False
                        )

                        df_hoja2.to_excel(
                            writer,
                            sheet_name="analisis_proveedor",
                            index=False
                        )

                    print("\nEXCEL MULTIHOJA ACTUALIZADO")
                    print(ruta_excel)

                    # =========================================
                    # ENVIAR CORREOS
                    # =========================================

                    emails_enviar = []

                    for key, value in datos_proveedor.items():

                        if "email_" in key:

                            if value and value != "NO ENCONTRADO":
                                emails_enviar.append(value)

                    emails_unicos = list(set(emails_enviar))

                    for correo in emails_unicos:

                        try:

                            print(f"\nENVIANDO CORREO A: {correo}")

                            enviar_correo(
                                correo,
                                proveedor
                            )

                            time.sleep(10)

                        except Exception as e:

                            print("ERROR ENVIO")
                            print(e)

                except Exception as e:

                    print("\nERROR FILA")

                    print(e)

                    try:
                        page_datos.close()
                    except:
                        pass

                    continue

        except Exception as e:

            print("\nERROR GENERAL")

            print(e)

        browser.close()


